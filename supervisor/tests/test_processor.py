"""Integration tests for the full processing pipeline."""

import tempfile
from pathlib import Path
import openpyxl
from openpyxl import Workbook
from app.processor import process_file


def _make_input(tmpdir: Path, builder) -> Path:
    """Helper: build a workbook with `builder(wb)`, save to tmpdir, return path."""
    wb = Workbook()
    builder(wb)
    in_path = tmpdir / "in.xlsx"
    wb.save(in_path)
    return in_path


def test_process_clean_file_produces_no_issues():
    with tempfile.TemporaryDirectory() as td:
        td = Path(td)
        def build(wb):
            ws = wb.active
            ws["A1"] = "name"; ws["B1"] = 42
        in_path = _make_input(td, build)
        out_path = td / "out.xlsx"
        result = process_file(in_path, out_path)
        assert out_path.exists()
        assert len(result.issues) == 0


def test_process_file_with_errors_includes_review_tab():
    with tempfile.TemporaryDirectory() as td:
        td = Path(td)
        def build(wb):
            ws = wb.active
            ws.title = "Data"
            ws["A1"] = "ok"
            ws["B1"] = "#REF!"
        in_path = _make_input(td, build)
        out_path = td / "out.xlsx"
        result = process_file(in_path, out_path)
        assert len(result.issues) >= 1

        # Reopen output and confirm the review tab is present
        out_wb = openpyxl.load_workbook(out_path)
        assert "_xlsx-for-ai" in out_wb.sheetnames
        assert "Data" in out_wb.sheetnames


def test_process_preserves_user_sheet_values():
    with tempfile.TemporaryDirectory() as td:
        td = Path(td)
        def build(wb):
            ws = wb.active
            ws.title = "Sales"
            ws["A1"] = "Region"; ws["B1"] = "Total"
            ws["A2"] = "North"; ws["B2"] = 1000
            ws["A3"] = "South"; ws["B3"] = 1500
        in_path = _make_input(td, build)
        out_path = td / "out.xlsx"
        process_file(in_path, out_path)
        out_wb = openpyxl.load_workbook(out_path)
        ws = out_wb["Sales"]
        assert ws["A2"].value == "North"
        assert ws["B2"].value == 1000


def test_process_strips_existing_review_tab():
    """If the input already has a _xlsx-for-ai tab, don't carry it over."""
    with tempfile.TemporaryDirectory() as td:
        td = Path(td)
        def build(wb):
            ws = wb.active
            ws.title = "Data"
            ws["A1"] = "value"
            ws["B1"] = 1
            wb.create_sheet("_xlsx-for-ai")["A1"] = "old review"
        in_path = _make_input(td, build)
        out_path = td / "out.xlsx"
        process_file(in_path, out_path)
        out_wb = openpyxl.load_workbook(out_path)
        # Output has its own (fresh) review tab, not the original "old review" content
        assert "_xlsx-for-ai" in out_wb.sheetnames
        review_ws = out_wb["_xlsx-for-ai"]
        # First cell should be the new title, not "old review"
        assert review_ws["A1"].value != "old review"


def test_process_review_tab_has_summary_when_clean():
    with tempfile.TemporaryDirectory() as td:
        td = Path(td)
        def build(wb):
            ws = wb.active
            ws["A1"] = 1
        in_path = _make_input(td, build)
        out_path = td / "out.xlsx"
        result = process_file(in_path, out_path)
        assert len(result.issues) == 0
        out_wb = openpyxl.load_workbook(out_path)
        review_ws = out_wb["_xlsx-for-ai"]
        # Should have the "no issues" summary somewhere
        all_text = []
        for row in review_ws.iter_rows(values_only=True):
            for v in row:
                if v:
                    all_text.append(str(v))
        text = " ".join(all_text)
        assert "No structural issues" in text or "no issues" in text.lower()
