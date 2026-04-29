"""End-to-end API tests using FastAPI's test client."""

import io
import openpyxl
from openpyxl import Workbook
from fastapi.testclient import TestClient

from app.main import app


client = TestClient(app)


def _xlsx_bytes(builder) -> bytes:
    wb = Workbook()
    builder(wb)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_health_endpoint():
    r = client.get("/health")
    assert r.status_code == 200
    assert r.json() == {"status": "ok"}


def test_index_serves_html():
    r = client.get("/")
    assert r.status_code == 200
    assert "xlsx-supervisor" in r.text


def test_upload_processes_clean_file():
    def build(wb):
        ws = wb.active
        ws["A1"] = "name"
        ws["B1"] = 42
    contents = _xlsx_bytes(build)
    r = client.post(
        "/upload",
        files={"file": ("clean.xlsx", contents, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 200
    body = r.json()
    assert body["filename"] == "clean.xlsx"
    assert body["issue_count"] == 0
    assert body["download_url"].startswith("/download/")


def test_upload_flags_errors():
    def build(wb):
        ws = wb.active
        ws["A1"] = "#REF!"
        ws["B1"] = "#DIV/0!"
    contents = _xlsx_bytes(build)
    r = client.post("/upload", files={"file": ("bad.xlsx", contents, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r.status_code == 200
    body = r.json()
    assert body["issue_count"] >= 2
    assert "formula_error" in body["by_type"]


def test_upload_rejects_non_xlsx():
    r = client.post("/upload", files={"file": ("data.csv", b"a,b,c\n1,2,3", "text/csv")})
    assert r.status_code == 400


def test_upload_rejects_empty():
    r = client.post("/upload", files={"file": ("empty.xlsx", b"", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    assert r.status_code == 400


def test_download_returns_processed_file():
    def build(wb):
        wb.active["A1"] = 42
    contents = _xlsx_bytes(build)
    r = client.post("/upload", files={"file": ("test.xlsx", contents, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")})
    file_id = r.json()["file_id"]
    dl = client.get(f"/download/{file_id}")
    assert dl.status_code == 200
    # Content should be a valid xlsx
    out_wb = openpyxl.load_workbook(io.BytesIO(dl.content))
    assert "_xlsx-for-ai" in out_wb.sheetnames


def test_download_unknown_id_404s():
    r = client.get("/download/nonexistent-id")
    assert r.status_code == 404


def test_upload_form_returns_html_fragment():
    def build(wb):
        wb.active["A1"] = 1
    contents = _xlsx_bytes(build)
    r = client.post(
        "/upload-form",
        files={"file": ("test.xlsx", contents, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert r.status_code == 200
    assert "Review complete" in r.text
    assert "Download" in r.text
