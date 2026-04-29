"""Unit tests for the reviewer's deterministic structural checks."""

import openpyxl
from openpyxl import Workbook
from app.reviewer import review_workbook


def test_clean_workbook_returns_no_issues():
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    ws["A1"] = "name"
    ws["B1"] = "value"
    ws["A2"] = "Alice"
    ws["B2"] = 100
    result = review_workbook(wb)
    assert len(result.issues) == 0


def test_formula_error_cell_is_flagged():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "#REF!"
    result = review_workbook(wb)
    assert len(result.issues) == 1
    assert result.issues[0].type == "formula_error"
    assert "#REF!" in result.issues[0].title


def test_multiple_error_types_each_flagged():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "#REF!"
    ws["B1"] = "#NAME?"
    ws["C1"] = "#DIV/0!"
    ws["D1"] = "normal text"  # not flagged
    result = review_workbook(wb)
    assert len(result.issues) == 3
    types = {i.title for i in result.issues}
    assert "Formula error: #REF!" in types
    assert "Formula error: #NAME?" in types
    assert "Formula error: #DIV/0!" in types


def test_broken_sheet_ref_in_formula_is_flagged():
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "=Detail!B5"  # 'Detail' sheet doesn't exist
    result = review_workbook(wb)
    flagged = [i for i in result.issues if i.type == "broken_sheet_ref"]
    assert len(flagged) == 1
    assert "Detail" in flagged[0].title


def test_valid_cross_sheet_ref_is_not_flagged():
    wb = Workbook()
    ws_a = wb.active
    ws_a.title = "Summary"
    ws_a["A1"] = "=Detail!B5"
    wb.create_sheet("Detail")
    result = review_workbook(wb)
    assert not any(i.type == "broken_sheet_ref" for i in result.issues)


def test_quoted_sheet_ref_handled():
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = "='Sales 2025'!B5"  # quoted sheet name
    result = review_workbook(wb)
    flagged = [i for i in result.issues if i.type == "broken_sheet_ref"]
    assert len(flagged) == 1


def test_hidden_row_with_data_is_flagged():
    wb = Workbook()
    ws = wb.active
    ws["A2"] = "secret"
    ws.row_dimensions[2].hidden = True
    result = review_workbook(wb)
    flagged = [i for i in result.issues if i.type == "hidden_row_with_data"]
    assert len(flagged) == 1


def test_hidden_empty_row_is_not_flagged():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "visible"
    ws.row_dimensions[2].hidden = True  # row 2 has no data
    result = review_workbook(wb)
    assert not any(i.type == "hidden_row_with_data" for i in result.issues)


def test_review_summary_counts():
    wb = Workbook()
    ws = wb.active
    ws["A1"] = "#REF!"
    ws["B1"] = "#NAME?"
    ws["C1"] = "#REF!"
    result = review_workbook(wb)
    assert result.summary_counts.get("formula_error") == 3


def test_skips_existing_review_tab_if_present():
    """Reviewing a previously-supervised file shouldn't re-flag its review tab."""
    wb = Workbook()
    ws_data = wb.active
    ws_data.title = "Data"
    ws_data["A1"] = 1
    review_ws = wb.create_sheet("_xlsx-for-ai")
    review_ws["A1"] = "#REF!"  # would normally be flagged
    result = review_workbook(wb)
    # The error in the review tab should not count
    assert not any("#REF!" in i.title for i in result.issues if i.sheet == "_xlsx-for-ai")
