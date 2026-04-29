"""File processing pipeline.

Takes an uploaded xlsx, runs the reviewer against it, and writes a new xlsx
that is the original file plus a `_xlsx-for-ai` review tab summarizing what
the reviewer found.

Architecture note: openpyxl handles the read; xlsxwriter handles the write.
We don't preserve the source file byte-for-byte (xlsxwriter doesn't read
xlsx). Instead we copy data + the review tab into a fresh file. This loses
some advanced formatting we don't need to round-trip in v1 — charts, pivots,
embedded images, etc. — and trades that loss for clean output and the
ability to add the review tab cleanly.

When/if we need full round-trip fidelity (preserve the source file plus add
the review tab), the path is openpyxl-only: load → modify → save adds the
review tab in-place. Tradeoff: openpyxl's fidelity quirks (it expands shared
formulas on read, which inflates file size for formula-heavy workbooks).
We'll evaluate that swap when a real user asks for the round-trip case.
"""

from datetime import datetime, UTC
from pathlib import Path
from openpyxl import load_workbook
import xlsxwriter

from .reviewer import review_workbook, ReviewResult, Issue


# Review-block content per issue type. Mirrors the supervisor framing the OSS
# CLI uses — review tab is a code-reviewer-style note, not just a warning list.
_REVIEW_TEMPLATES = {
    "formula_error": {
        "what_happened": "Excel returned an error literal in this cell. The formula tried to compute something but couldn't.",
        "what_we_did": "Flagged the cell. Did not attempt to repair the formula automatically — there are too many ways a formula can break, and silently changing it could hide a real problem.",
        "risk": "If you ship this file, anyone who opens it will see the same error. Downstream cells that depend on it may also break or compute wrong values.",
        "tradeoff": "Manually fixing it requires understanding the original intent. Auto-fixing it would risk getting the answer wrong silently.",
        "alternative": "Open the cell in Excel, review the formula, and either repair the reference or replace with a hardcoded value if the source is unrecoverable.",
    },
    "broken_sheet_ref": {
        "what_happened": "A formula references a sheet name that doesn't exist in this workbook.",
        "what_we_did": "Flagged the cells. Could not infer the intended sheet name without guessing.",
        "risk": "Excel will show #REF! errors for these cells. Anything dependent on them will also break.",
        "tradeoff": "If we auto-renamed a sheet to match, we might rename the wrong one or break unrelated formulas.",
        "alternative": "Either restore the missing sheet (rename an existing one, or add a new one with the expected name) or update the formula to point at an existing sheet.",
    },
    "hidden_row_with_data": {
        "what_happened": "A row was hidden in the original file but contains data.",
        "what_we_did": "Flagged the row. Did not unhide or delete it.",
        "risk": "Hidden data can silently feed into totals or formulas without the reader realizing. It can also represent in-progress or scratch work the original author meant to remove.",
        "tradeoff": "Unhiding it changes the visual layout (which may surprise the reader). Deleting it loses data that might be needed.",
        "alternative": "Decide whether the data should be visible (unhide) or removed (delete). If it's intentional auxiliary data, add a cell comment documenting why.",
    },
    "hidden_col_with_data": {
        "what_happened": "A column was hidden but contains data.",
        "what_we_did": "Flagged the column. Did not unhide.",
        "risk": "Same as hidden rows — silent contribution to totals; silent on-screen values.",
        "tradeoff": "Same tradeoff.",
        "alternative": "Unhide if it should be visible, delete if not needed, or document why it's hidden.",
    },
    "external_link": {
        "what_happened": "The workbook contains a link to another external workbook.",
        "what_we_did": "Flagged the link.",
        "risk": "External links break when the file moves, gets emailed, or is opened on a different machine. Excel will prompt the recipient about updating links and may show #REF! errors when the target isn't found.",
        "tradeoff": "Removing the link severs the connection; if the target is updated, this file won't reflect changes.",
        "alternative": "If you need the latest values: paste them in as values via Edit > Paste Special > Values. If you don't need the link: remove it via Excel's Data > Edit Links.",
    },
}


def _fallback_template(issue: Issue) -> dict:
    return {
        "what_happened": issue.description,
        "what_we_did": "Flagged the cell.",
        "risk": "Unknown — review manually.",
        "tradeoff": "Unknown.",
        "alternative": issue.suggestion,
    }


def process_file(input_path: Path, output_path: Path) -> ReviewResult:
    """Process a single xlsx: read, review, write a copy with the review tab."""
    # Load source with openpyxl (preserves formulas + values)
    src_wb = load_workbook(input_path, data_only=False)
    review = review_workbook(src_wb)

    # Build output with xlsxwriter
    out_wb = xlsxwriter.Workbook(str(output_path))

    # Copy each user sheet
    for src_ws in src_wb.worksheets:
        if src_ws.title == "_xlsx-for-ai":
            continue  # don't carry over a previous review tab
        out_ws = out_wb.add_worksheet(src_ws.title)
        for row in src_ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                v = cell.value
                # Formulas
                if isinstance(v, str) and v.startswith("="):
                    out_ws.write_formula(cell.row - 1, cell.column - 1, v)
                else:
                    try:
                        out_ws.write(cell.row - 1, cell.column - 1, v)
                    except Exception:
                        out_ws.write(cell.row - 1, cell.column - 1, str(v))

    # Add the review tab
    _add_review_tab(out_wb, review)

    out_wb.close()
    return review


def _add_review_tab(out_wb, review: ReviewResult) -> None:
    """Add the `_xlsx-for-ai` review tab to the output workbook."""
    ws = out_wb.add_worksheet("_xlsx-for-ai")

    bold = out_wb.add_format({"bold": True, "font_size": 14})
    bold_small = out_wb.add_format({"bold": True})
    italic = out_wb.add_format({"italic": True, "font_color": "#666666"})
    section_header = out_wb.add_format({"bold": True, "font_size": 12, "bg_color": "#E7F0F8"})
    wrap = out_wb.add_format({"text_wrap": True, "valign": "top"})
    wrap_bold = out_wb.add_format({"text_wrap": True, "valign": "top", "bold": True})
    table_header = out_wb.add_format({"bold": True, "bg_color": "#EEEEEE"})

    ws.set_column(0, 0, 18)
    ws.set_column(1, 1, 12)
    ws.set_column(2, 2, 18)
    ws.set_column(3, 3, 80)

    # Header
    ws.merge_range("A1:D1", "xlsx-supervisor review report", bold)
    ws.merge_range("A2:D2", f"Generated {datetime.now(UTC).strftime('%Y-%m-%d %H:%M:%S')} UTC", italic)
    ws.merge_range(
        "A3:D3",
        "This file was reviewed by xlsx-supervisor. Below: what we found, why "
        "it matters, what could go wrong, and how to override our recommendations "
        "if you disagree. Cell values are unchanged from your upload — these "
        "notes describe issues we want you to look at.",
        wrap,
    )
    ws.set_row(2, 60)

    if not review.issues:
        ws.write("A5", "No structural issues detected.", section_header)
        ws.merge_range("A5:D5", "No structural issues detected.", section_header)
        return

    # Group issues by type
    by_type: dict[str, list[Issue]] = {}
    for issue in review.issues:
        by_type.setdefault(issue.type, []).append(issue)

    r = 4  # 0-indexed; row 5 in spreadsheet
    for issue_type, group in by_type.items():
        first = group[0]
        # Section heading
        ws.merge_range(r, 0, r, 3, f"Issue: {first.title}  ({len(group)} occurrence{'s' if len(group) > 1 else ''})", section_header)
        r += 1

        templ = _REVIEW_TEMPLATES.get(issue_type) or _fallback_template(first)

        for label, key in [
            ("What happened", "what_happened"),
            ("What we did", "what_we_did"),
            ("Risk", "risk"),
            ("Tradeoff", "tradeoff"),
            ("Alternative", "alternative"),
        ]:
            ws.write(r, 0, label, wrap_bold)
            ws.merge_range(r, 1, r, 3, templ.get(key, ""), wrap)
            ws.set_row(r, 40)
            r += 1

        # Affected cells (compact list)
        cells = [f"{i.sheet}!{i.cell}" for i in group]
        if len(cells) <= 10:
            cell_summary = ", ".join(cells)
        else:
            cell_summary = ", ".join(cells[:10]) + f", ... and {len(cells) - 10} more (full list at the bottom)"
        ws.write(r, 0, "Affected", wrap_bold)
        ws.merge_range(r, 1, r, 3, cell_summary, wrap)
        ws.set_row(r, 30)
        r += 1
        r += 1  # spacer

    # Detail table
    ws.merge_range(r, 0, r, 3, "Full detail (one row per affected cell)", bold_small)
    r += 1
    for col_idx, header in enumerate(["Sheet", "Cell", "Severity", "Description"]):
        ws.write(r, col_idx, header, table_header)
    r += 1
    for issue in review.issues:
        ws.write(r, 0, issue.sheet)
        ws.write(r, 1, issue.cell)
        ws.write(r, 2, issue.severity)
        ws.write(r, 3, issue.description, wrap)
        r += 1
