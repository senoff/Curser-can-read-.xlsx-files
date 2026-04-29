"""Generate a review of a workbook.

For the MVP, the "review" is a set of deterministic structural checks — no
LLM call. This is intentional:

  - Lets us ship the upload→process→download loop without committing to a
    BYO-AI vs we-pay-for-the-key decision yet
  - Structural checks (formula errors, broken refs, hidden data) catch the
    majority of bookkeeping mistakes that humans want flagged
  - Cheap and deterministic — no external API, no rate limits, no costs

The reviewer's output is a list of Issue objects. The processor turns each
into a section of the `_xlsx-for-ai` review tab embedded in the output file
— same shape as the CLI's write-mode review tab, so the user sees consistent
output whether the file was processed by the OSS CLI or the supervisor.

Future: when we lock in the LLM strategy, add a second `LLMReviewer` class
implementing the same Issue-emitting interface; the processor doesn't need
to change.
"""

from dataclasses import dataclass, field
from typing import Any
from openpyxl import Workbook
from openpyxl.utils import get_column_letter


# ---------------------------------------------------------------------------
# Issue model
# ---------------------------------------------------------------------------

@dataclass
class Issue:
    """A single thing the reviewer found."""
    type: str           # short machine-readable category
    title: str          # human-readable section heading for the report tab
    sheet: str
    cell: str           # e.g. "A1" or "A1:D5" for ranges
    severity: str       # "error" | "warning" | "info"
    description: str    # what's wrong
    suggestion: str     # recommended fix


@dataclass
class ReviewResult:
    """Output of a review pass."""
    issues: list[Issue] = field(default_factory=list)
    summary_counts: dict[str, int] = field(default_factory=dict)

    def add(self, issue: Issue) -> None:
        self.issues.append(issue)
        self.summary_counts[issue.type] = self.summary_counts.get(issue.type, 0) + 1


# ---------------------------------------------------------------------------
# Deterministic structural checks
# ---------------------------------------------------------------------------

# Values openpyxl returns for the common Excel error types
_EXCEL_ERROR_VALUES = {"#REF!", "#NAME?", "#DIV/0!", "#VALUE!", "#NULL!", "#NUM!", "#N/A"}


def _check_error_cells(ws, result: ReviewResult) -> None:
    """Flag cells whose value is one of Excel's standard error literals."""
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str) and v in _EXCEL_ERROR_VALUES:
                result.add(Issue(
                    type="formula_error",
                    title=f"Formula error: {v}",
                    sheet=ws.title,
                    cell=cell.coordinate,
                    severity="error",
                    description=f"Cell {cell.coordinate} on '{ws.title}' contains a {v} error. The formula isn't computing — usually because it references a deleted sheet, a renamed range, or has a type mismatch.",
                    suggestion=f"Open the cell and check the formula. Common fixes: re-create the missing reference, repair the named range, or replace the formula with a hardcoded value if the source data isn't recoverable.",
                ))


def _check_broken_formula_refs(ws, sheet_names: set[str], result: ReviewResult) -> None:
    """Flag formulas that reference sheet names not present in the workbook."""
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if not isinstance(v, str) or not v.startswith("="):
                continue
            # Find sheet refs in the formula: 'SheetName'! or SheetName!
            import re
            refs = re.findall(r"(?:'([^']+)'|([A-Za-z_][A-Za-z0-9_ ]*))!", v)
            for quoted, plain in refs:
                ref_name = quoted or plain
                if not ref_name:
                    continue
                ref_name = ref_name.strip()
                if ref_name not in sheet_names:
                    result.add(Issue(
                        type="broken_sheet_ref",
                        title=f"Formula references missing sheet: '{ref_name}'",
                        sheet=ws.title,
                        cell=cell.coordinate,
                        severity="error",
                        description=f"The formula in {cell.coordinate} on '{ws.title}' references a sheet called '{ref_name}', but that sheet doesn't exist in this workbook. The formula will return #REF! when Excel opens the file.",
                        suggestion=f"Either rename a sheet to '{ref_name}', update the formula to point at an existing sheet, or replace the formula with the value it should have produced.",
                    ))


def _check_hidden_data(ws, result: ReviewResult) -> None:
    """Flag hidden rows or columns that contain data — common source of confusion."""
    # Hidden rows with values
    for row in ws.iter_rows():
        if not row:
            continue
        row_dim = ws.row_dimensions.get(row[0].row)
        if row_dim and row_dim.hidden:
            for cell in row:
                if cell.value is not None and cell.value != "":
                    result.add(Issue(
                        type="hidden_row_with_data",
                        title=f"Hidden row {cell.row} contains data",
                        sheet=ws.title,
                        cell=f"row {cell.row}",
                        severity="warning",
                        description=f"Row {cell.row} on '{ws.title}' is hidden but contains values. Hidden data often indicates rows the original author meant to remove or rows that are silently affecting totals/formulas elsewhere.",
                        suggestion="Either unhide the row (so it's visible to the reader) or delete it (if it's not needed). If it's intentionally hidden as auxiliary calculation data, document that in a comment.",
                    ))
                    break  # one issue per hidden row
    # Hidden columns with values
    for col_letter, col_dim in (ws.column_dimensions or {}).items():
        if not col_dim.hidden:
            continue
        try:
            col_idx = ws[col_letter + "1"].column
        except Exception:
            continue
        # Walk a sample of cells in this column for non-empty values
        for r in range(1, min(ws.max_row, 100) + 1):
            cell = ws.cell(row=r, column=col_idx)
            if cell.value is not None and cell.value != "":
                result.add(Issue(
                    type="hidden_col_with_data",
                    title=f"Hidden column {col_letter} contains data",
                    sheet=ws.title,
                    cell=f"col {col_letter}",
                    severity="warning",
                    description=f"Column {col_letter} on '{ws.title}' is hidden but contains values. Same risk as hidden rows — silent contribution to totals or formulas.",
                    suggestion="Unhide the column or remove it. If it's an intentional helper column, document why.",
                ))
                break


def _check_external_links(wb, result: ReviewResult) -> None:
    """Flag external workbook links, which break when the file moves."""
    try:
        external_links = getattr(wb, "_external_links", None) or []
        for link in external_links:
            target = getattr(link, "file_link", None) or getattr(link, "Target", "<unknown>")
            result.add(Issue(
                type="external_link",
                title=f"External link to another workbook",
                sheet="(workbook-level)",
                cell="-",
                severity="warning",
                description=f"This workbook links to an external file: {target}. External links commonly break when the file moves, gets emailed, or is opened on another machine.",
                suggestion="If the linked data is needed, paste it as values into this workbook. If the link is no longer needed, remove it via Excel's Edit Links dialog.",
            ))
    except Exception:
        pass


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def review_workbook(wb: Workbook) -> ReviewResult:
    """Run the full set of structural checks against a workbook."""
    result = ReviewResult()
    sheet_names = {s.title for s in wb.worksheets}
    for ws in wb.worksheets:
        # Skip our own report tab if re-reviewing a previously-supervised file
        if ws.title == "_xlsx-for-ai":
            continue
        _check_error_cells(ws, result)
        _check_broken_formula_refs(ws, sheet_names, result)
        _check_hidden_data(ws, result)
    _check_external_links(wb, result)
    return result
